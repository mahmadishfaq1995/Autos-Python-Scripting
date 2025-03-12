from openpyxl import load_workbook
try:
    try:
        wb = load_workbook('/Users/ahmed/PycharmProjects/QB/log_failure.xlsx')
        sheet_name = wb['Failed_Test_Cases']
        wb.remove(sheet_name)
        wb.save("log_failure.xlsx")
    except:
        print("Log_failure file does not exist")
except:
    print("Unexpected Error in clearing Log_failure.xlsx file")
try:
    try:
        wb = load_workbook('/Users/ahmed/PycharmProjects/QB/log_failure.xlsx')
        sheet_name = wb['Success_Test_Cases']
        wb.remove(sheet_name)
        wb.save("log_success.xlsx")
    except:
        print("Log_success file does not exist")
except:
    print("Unexpected Error in clearing Log_success.xlsx file")



