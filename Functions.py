from Login import *
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime
# tenant name here
tenant_name = "OLX/AUTOS"


# url of project here
def url():
    #olx
    urrl: str = 'https://staging-olx.empgautos.com/admin'
    #Dubizzle
    #urrl: str = 'https://staging-dubizzle.empgautos.com/admin/login'
    return urrl


# logging function
def log_success(value):
    print(value)
    from openpyxl import Workbook
    from openpyxl import load_workbook
    try:
        filepath = '/Users/Muhammad.Ahmad/PycharmProjects/AUTOS/'
        filename = 'log_success.xlsx'
        wb = load_workbook(filepath + filename)
    except:
        wb = Workbook()
        wb.save(filepath + filename)
        wb = load_workbook(filepath + filename)
        sheet = wb['Sheet']
        sheet.title = "Success_Test_Cases"
        wb.save(filepath + filename)
    ws = wb.active
    ws.cell(row=1, column=1).value = "Test Case Description"
    start_row = 2
    start_column = 1
    for i in value:
        if ws.cell(row=start_row, column=start_column).value is None:
            # print(start_row)
            ws.cell(row=start_row, column=start_column).value = value
            ws.column_dimensions["A"].width = 100
            wb.save(filename)
            break


        else:
            start_row += 1
            # print("in else", start_row)
            continue


def log_failure(value) -> object:
    print(value)
    from openpyxl import Workbook
    from openpyxl import load_workbook
    try:
        filepath = '/Users/Muhammad.Ahmad/PycharmProjects/AUTOS/'
        filename = 'Log_failure.xlsx'
        wb = load_workbook(filepath + filename)
    except:
        wb = Workbook()
        wb.save(filepath + filename)
        wb = load_workbook(filepath + filename)
        sheet = wb['Sheet']
        sheet.title = "Failure_Test_Cases"
        wb.save(filepath + filename)

        # if wb['Success_Test_Cases'] is True:
        #     pass
        # else:
        #     sheet = wb['Sheet']
        #     sheet.title = "Success_Test_Cases"
    ws = wb.active
    ws.cell(row=1, column=1).value = "Test Case Description"
    start_row = 2
    start_column = 1
    for i in value:
        if ws.cell(row=start_row, column=start_column).value is None:
            # print(start_row)
            ws.cell(row=start_row, column=start_column).value = value
            ws.column_dimensions["A"].width = 100
            wb.save(filename)
            break


        else:
            start_row += 1
            # print("in else", start_row)
            continue


# logout function

def logout(driver, email="", password=""):
    try:
        elem = driver.find_element_by_xpath(
            "/html/body/div/div/div/div[1]/div/div[1]/div/div[3]/ul/li[3]/div/div[1]/div/div/span/span")
        hover = ActionChains(driver).move_to_element(elem)
        hover.perform()
        elem = driver.find_element_by_link_text("Log out")
        elem.click()
        time.sleep(5)
        try:
            elem = driver.find_element_by_xpath("//span[contains(text(),'Login')]")
            if "Login" in elem.text:
                log_success(tenant_name + " ""Logput  >> Pass")
                log_success(tenant_name + " ""Logput  >> User Redirected to Login Page >> Pass")
        except:
            log_failure(tenant_name + " ""Logput  >> User Redirected to Login Page >> issue in "
                                      "verifying Logout")
    except:
        log_success(tenant_name + " ""Logput  >> Unable to Locate Logout Link.")

def clearall(driver,Pagename="",Fieldname=""):
    try:
        time.sleep(2)
        elem = driver.find_element_by_xpath(
            '//span[@class][contains(text(),"Clear All Filters")]')
        elem.click()
        log_success(
            tenant_name + " ""Login >> "+Pagename+" Page >>  Search by "+Fieldname+" >> Clear Filter >> Pass")
    except:
        log_failure(
            tenant_name + " ""Login >> " + Pagename + " Page >>  Search by " + Fieldname + " >> Clear Filter >> Fail")


def home(driver,pagename=""):
    try:
        time.sleep(2)
        elem = driver.find_element_by_xpath("//li[@class='breadcrumb-item']/a")
        elem.click()
        try:
            elem = driver.find_element_by_xpath("//li[@class='breadcrumb-item active'][text()='Dashboard']")
            if elem:
                log_success(
                    tenant_name + " ""Login >> " + Pagename + " Page >>  Home icon Redirection  >> Pass")
            else:
                log_failure(
                    tenant_name + " ""Login >> " + Pagename + " Page >>  Home icon Redirection  >> Fail")
        except:
            log_failure(
                tenant_name + " ""Login >> " + Pagename + " Page >>  Home icon Redirection  >> Unexpected Error")
    except:
        log_failure(
            tenant_name + " ""Login >> " + Pagename + " Page >>  Home icon Redirection >> Element not found")

